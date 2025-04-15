import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import re
import os
import sys
import openpyxl

class DateHandler:
    @staticmethod
    def parse_date(date_str):
        """
        Parse date string in various formats and convert to datetime object.
        Handles DD-MM-YY, D-M-YY, DD-MM-YYYY formats and the Y2K problem.
        """
        if not date_str or date_str.strip() == "":
            return None
            
        date_str = date_str.strip()
        
        # Regular expressions for various date formats
        patterns = [
            r'^(\d{1,2})[/-](\d{1,2})[/-](\d{2})$',  # DD-MM-YY or D-M-YY
            r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$',  # DD-MM-YYYY or D-M-YYYY
        ]
        
        for pattern in patterns:
            match = re.match(pattern, date_str)
            if match:
                day, month, year = match.groups()
                
                # Handle Y2K problem for two-digit years
                if len(year) == 2:
                    year_int = int(year)
                    if year_int >= 50:  # Assume 1950-1999
                        year = f"19{year}"
                    else:  # Assume 2000-2049
                        year = f"20{year}"
                
                try:
                    return datetime(int(year), int(month), int(day))
                except ValueError:
                    raise ValueError(f"Invalid date: {date_str}")
        
        raise ValueError(f"Unsupported date format: {date_str}")
    
    @staticmethod
    def format_date(date_obj):
        """Format datetime object to DD-MM-YY string"""
        if date_obj is None:
            return ""
        return date_obj.strftime("%d-%m-%y")
    
    @staticmethod
    def get_increment_date(appointment_date):
        """
        Determine increment date based on appointment month:
        - Jan-Jun appointments → January increments (next year)
        - Jul-Dec appointments → July increments (next year)
        """
        # Next year's increment date
        next_year = appointment_date.year + 1
        
        if appointment_date.month <= 6:
            # Jan-Jun: increment on January 1st next year
            return datetime(next_year, 1, 1)
        else:
            # Jul-Dec: increment on July 1st next year
            return datetime(next_year, 7, 1)
    
    @staticmethod
    def generate_increment_dates(start_date, end_date=datetime(2007, 3, 31)):
        """Generate all increment dates between start date and end date (March 31, 2007)"""
        if start_date is None or end_date is None:
            return []
            
        increment_dates = []
        
        # Get first increment date (next year)
        current_date = DateHandler.get_increment_date(start_date)
        
        # Generate subsequent increment dates
        while current_date <= end_date:
            increment_dates.append(current_date)
            
            # Add one year for the next increment
            if current_date.month == 1:
                current_date = datetime(current_date.year + 1, 1, 1)
            else:  # month == 7
                current_date = datetime(current_date.year + 1, 7, 1)
        
        return increment_dates


class ExcelDataHandler:
    def __init__(self, excel_path=None):
        self.excel_path = excel_path
        self.salary_tables = {}
        
    def load_excel_file(self, file_path=None):
        """Load and parse the Excel workbook"""
        if file_path:
            self.excel_path = file_path
        
        if not self.excel_path or not os.path.exists(self.excel_path):
            raise FileNotFoundError("Excel file not found")
        
        try:
            xl = pd.ExcelFile(self.excel_path)
            self._parse_sheets(xl)
            return True
        except Exception as e:
            raise Exception(f"Error loading Excel file: {str(e)}")
    
    def _parse_sheets(self, excel_file):
        """Parse each sheet in the Excel workbook"""
        for sheet_name in excel_file.sheet_names:
            # Skip sheets that don't match the expected format
            if not self._is_valid_sheet_name(sheet_name) and not sheet_name.upper() in ['HSS', 'HATISS']:
                continue
            
            df = pd.read_excel(excel_file, sheet_name)
            self._process_salary_table(sheet_name, df)
    
    def _is_valid_sheet_name(self, sheet_name):
        """Validate sheet name format (YYYY-MM-DD_to_YYYY-MM-DD)"""
        pattern = r'^\d{4}-\d{2}-\d{2}_to_\d{4}-\d{2}-\d{2}$'
        return re.match(pattern, sheet_name) is not None
    
    def _process_salary_table(self, sheet_name, df):
        """
        Extract salary data from dataframe with the following structure:
        - First row (0): Step labels (1-15)
        - First column (A): Grade level labels (1-17)
        - Data cells (B1:P17): Salary values for each grade/step combination
        """
        table = {}

        try:
            # Get actual number of rows in the dataframe
            max_rows = min(17, len(df))
        
            # Start from row 0 for grade levels (instead of skipping the header row)
            for row_idx in range(0, max_rows):  # Process available rows (for grades 1-17)
                # The grade level corresponds to row_idx + 1 (since grade levels start from 1)
                grade_level = row_idx + 1
                if grade_level >= 11:
                    grade_level += 1  # Adjust for skipping grade 11
                
                # Now grade_level correctly corresponds to the actual grade (1,2,3,...,10,12,13,...,17)
                grade_data = {}

                # Get actual number of columns in the dataframe
                max_cols = min(16, df.shape[1])

                # Process available columns for steps (skip first column which has grade labels)
                for step_idx in range(1, max_cols):  # Columns B-P (indices 1-15)
                    if row_idx < len(df) and step_idx < df.shape[1]:  # Additional safety check
                        value = df.iloc[row_idx, step_idx]
                        if pd.notna(value):  # Check if value is not NaN
                            try:
                                grade_data[step_idx] = float(value)
                            except (ValueError, TypeError):
                                # Skip non-numeric values
                                continue
            
                if grade_data:  # Only add if we have valid data
                    table[grade_level] = grade_data

        except Exception as e:
            raise Exception(f"Error processing salary table: {str(e)}")

        # Store the table with period information
        if sheet_name.upper() in ['HSS', 'HATISS']:
            self.salary_tables[sheet_name.upper()] = table
        else:
            # Extract date range from sheet name
            start_date_str, end_date_str = sheet_name.split('_to_')
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d')

            self.salary_tables[(start_date, end_date)] = table

        return table
    

    def get_salary_table_for_date(self, date, table_type=None):
        """Get the salary table effective on a specific date"""
        if date is None:
            return None
    
        # Check special tables first
        if table_type:
            if table_type.upper() == 'HSS' and 'HSS' in self.salary_tables:
                # HSS was implemented from May 1, 2005
                if date >= datetime(2005, 5, 1):
                    return self.salary_tables['HSS']
            elif table_type.upper() == 'HATISS' and 'HATISS' in self.salary_tables:
                # HATISS was implemented from November 1, 2005
                if date >= datetime(2005, 11, 1):
                    return self.salary_tables['HATISS']
    
        # Find the regular table for the date
        for (start_date, end_date), table in self.salary_tables.items():
            if isinstance(start_date, datetime) and isinstance(end_date, datetime):
                if start_date <= date <= end_date:
                    return table

        return None
    
    def get_salary_value(self, date, grade_level, step, table_type=None):
        """Get the salary value for a specific date, grade level, and step"""
        table = self.get_salary_table_for_date(date, table_type)
        if not table:
            return None
            
        # Adjust grade level based on table type and date
        adjusted_grade = self._adjust_grade_level(grade_level, table_type, date)
        
        # Get salary value
        if adjusted_grade in table and step in table[adjusted_grade]:
            return table[adjusted_grade][step]
        
        return None
    
    def _adjust_grade_level(self, grade_level, table_type, date):
        """
        Adjust grade level based on HATISS/HSS implementation:
        - HATISS (from November 1, 2005):
            - Levels 1-10: Move back 1 grade
            - Levels 12-17: Move back 2 grades
        - HSS (from May 1, 2005):
            - All levels move back 1 grade
        """
        if table_type and table_type.upper() == 'HATISS' and date >= datetime(2005, 11, 1):
            if 1 <= grade_level <= 10:
                return grade_level - 1
            elif 12 <= grade_level <= 17:
                return grade_level - 2
        elif table_type and table_type.upper() == 'HSS' and date >= datetime(2005, 5, 1):
            return grade_level - 1
        
        return grade_level
    
    def find_equivalent_step(self, old_salary, new_grade, date, table_type=None):
        """
        Find the step in the new grade that gives a salary equal to or greater than the old salary
        """
        if old_salary is None:
            return 1
            
        table = self.get_salary_table_for_date(date, table_type)
        if not table:
            return 1
            
        adjusted_grade = self._adjust_grade_level(new_grade, table_type, date)
        if adjusted_grade not in table:
            return 1
            
        # Get all steps and salaries for the new grade
        grade_steps = table[adjusted_grade]
        
        # Find the first step with salary >= old_salary
        for step in sorted(grade_steps.keys()):
            if grade_steps[step] >= old_salary:
                return step
        
        # If no step has higher salary, return the highest step
        return max(grade_steps.keys())


class PromotionEntry:
    def __init__(self, date, promotion_type, new_grade, new_step=None):
        self.date = date
        self.promotion_type = promotion_type
        self.new_grade = new_grade
        # Handle step value carefully
        if new_step is not None and new_step != '':
            try:
                self.new_step = int(new_step)
            except (ValueError, TypeError):
                self.new_step = None
        else:
            self.new_step = None
    
    def __str__(self):
        date_str = DateHandler.format_date(self.date)
        return f"{date_str}: {self.promotion_type} to GL {self.new_grade}" + (f" Step {self.new_step}" if self.new_step is not None else "")

    def _add_promotion(self):
        """Add a promotion entry"""
        if not self._validate_session():
            return
            
        # Make sure the button is in "Add" mode
        self.add_promotion_button.config(text="Add Promotion", command=self._add_promotion)
        
        try:
            # Validate inputs
            date_str = self.promotion_date_var.get()
            promotion_type = self.promotion_type_var.get()
            new_grade = self.promotion_grade_var.get()
            
            # Handle step value carefully
            step_value = self.promotion_step_var.get().strip()
            new_step = None
            if step_value:
                try:
                    new_step = int(step_value)
                except (ValueError, TypeError):
                    new_step = None
            
            if not date_str or not promotion_type or not new_grade:
                messagebox.showerror("Error", "Please fill all required promotion details")
                return
                
            try:
                promotion_date = DateHandler.parse_date(date_str)
                if not promotion_date:
                    raise ValueError("Invalid date format")
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid date: {str(e)}")
                return
            
            # Validate promotion date is between appointment and end date
            appointment_date_str = self.appointment_date_var.get()
            if appointment_date_str:
                try:
                    appointment_date = DateHandler.parse_date(appointment_date_str)
                    if promotion_date < appointment_date:
                        messagebox.showerror("Error", "Promotion date cannot be before appointment date")
                        return
                    if promotion_date > datetime(2007, 3, 31):
                        messagebox.showerror("Error", "Promotion date cannot be after March 31, 2007")
                        return
                except ValueError:
                    pass  # Will be caught during calculation
            
            # Create promotion entry with optional step
            promotion = PromotionEntry(promotion_date, promotion_type, new_grade, new_step)
            
            # Add to treeview
            item_id = self.promotion_tree.insert('', 'end', values=(
                DateHandler.format_date(promotion_date),
                new_grade,
                new_step if new_step is not None else '',  # Show empty string if no step specified
                promotion_type
            ))
            
            # Store promotion with its treeview ID
            self.promotion_list.append({
                'id': item_id,
                'promotion': promotion
            })
            
            # Clear entry fields
            self.promotion_date_var.set('')
            self.promotion_grade_var.set('')
            self.promotion_step_var.set('')  # Clear step field
            
            self.status_var.set(f"Added {promotion_type} to Grade {new_grade}" + 
                             (f" Step {new_step}" if new_step is not None else ""))
            
            # Set focus back to the date entry field
            self.promotion_date_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add promotion: {str(e)}")

    def _update_promotion(self, idx, item_id):
        """Update an existing promotion with new values"""
        try:
            # Validate inputs
            date_str = self.promotion_date_var.get()
            promotion_type = self.promotion_type_var.get()
            new_grade = self.promotion_grade_var.get()
            
            # Handle step value carefully
            step_value = self.promotion_step_var.get().strip()
            new_step = None
            if step_value:
                try:
                    new_step = int(step_value)
                except (ValueError, TypeError):
                    new_step = None
            
            if not date_str or not promotion_type or not new_grade:
                messagebox.showerror("Error", "Please fill all required promotion details")
                return
                
            try:
                promotion_date = DateHandler.parse_date(date_str)
                if not promotion_date:
                    raise ValueError("Invalid date format")
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid date: {str(e)}")
                return
            
            # Validate promotion date is between appointment and end date
            appointment_date_str = self.appointment_date_var.get()
            if appointment_date_str:
                try:
                    appointment_date = DateHandler.parse_date(appointment_date_str)
                    if promotion_date < appointment_date:
                        messagebox.showerror("Error", "Promotion date cannot be before appointment date")
                        return
                    if promotion_date > datetime(2007, 3, 31):
                        messagebox.showerror("Error", "Promotion date cannot be after March 31, 2007")
                        return
                except ValueError:
                    pass  # Will be caught during calculation
            
            # Create updated promotion entry with optional step
            updated_promotion = PromotionEntry(promotion_date, promotion_type, new_grade, new_step)
            
            # Update the promotion in our list
            self.promotion_list[idx]['promotion'] = updated_promotion
            
            # Update the treeview
            self.promotion_tree.item(item_id, values=(
                DateHandler.format_date(promotion_date),
                new_grade,
                new_step if new_step is not None else '',  # Show empty string if no step specified
                promotion_type
            ))
            
            # Reset form and button
            self.promotion_date_var.set('')
            self.promotion_grade_var.set('')
            self.promotion_step_var.set('')  # Clear step field
            self.add_promotion_button.config(text="Add Promotion", command=self._add_promotion)
            
            self.status_var.set(f"Updated {promotion_type} to Grade {new_grade}" + 
                             (f" Step {new_step}" if new_step is not None else ""))
            
            # Set focus back to the date entry field
            self.promotion_date_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update promotion: {str(e)}")


class SalaryProgressionCalculator:
    def __init__(self, excel_handler):
        self.excel_handler = excel_handler
        self.unit_type = None
        self.sub_type = None
        self.appointment_date = None
        self.initial_grade = None
        self.initial_step = None
        self.promotions = []
        
    def set_parameters(self, unit_type, sub_type, appointment_date, initial_grade, initial_step):
        """Set the basic parameters for calculation"""
        self.unit_type = unit_type
        self.sub_type = sub_type if unit_type in ["Mainstream", "Tescom"] else None
        self.appointment_date = appointment_date
        self.initial_grade = initial_grade
        self.initial_step = initial_step
    
    def add_promotion(self, promotion):
        """Add a promotion entry"""
        self.promotions.append(promotion)
        # Sort promotions by date
        self.promotions.sort(key=lambda p: p.date)
    
    def remove_promotion(self, index):
        """Remove a promotion entry by index"""
        if 0 <= index < len(self.promotions):
            del self.promotions[index]
    
    def get_max_step_for_grade(self, grade):
        """
        Determine the maximum step based on grade level:
        - Grade 1-10: max step 15
        - Grade 12-14: max step 11
        - Grade 15-17: max step 9
        """
        if 1 <= grade <= 10:
            return 15
        elif 12 <= grade <= 14:
            return 11
        elif 15 <= grade <= 17:
            return 9
        else:
            raise ValueError(f"Invalid grade level: {grade}")
    
    def calculate_progression(self):
        """
        Calculate the salary progression up to March 31, 2007
        Returns a list of yearly records with grade, step, and date information
        """
        if not all([self.unit_type, self.appointment_date, self.initial_grade, self.initial_step]):
            raise ValueError("Missing required parameters for calculation")
        
        progression = []
        current_grade = self.initial_grade
        current_step = self.initial_step
        end_date = datetime(2007, 3, 31)
        
        # Start with appointment record
        progression.append({
            'date': self.appointment_date,
            'grade': current_grade,
            'step': current_step,
            'event': 'Appointment',
            'salary': self.excel_handler.get_salary_value(
                self.appointment_date, current_grade, current_step, self.sub_type
            )
        })
        
        # Generate first increment date (no increment in appointment year)
        next_increment_date = DateHandler.get_increment_date(self.appointment_date)
        last_increment_date = None  # Track the last increment date
        
        # Process events chronologically until end date
        while next_increment_date <= end_date:
            # Get current state from last progression record
            current_record = progression[-1]
            current_grade = current_record['grade']
            current_step = current_record['step']
            
            # Get maximum step for current grade
            max_step = self.get_max_step_for_grade(current_grade)
            
            # Process promotions that occur before this increment
            promotions_before_increment = [p for p in self.promotions 
                                        if p.date > progression[-1]['date'] and p.date < next_increment_date]
            
            for promotion in sorted(promotions_before_increment, key=lambda p: p.date):
                # Get promotion details
                new_grade = promotion.new_grade
                promotion_date = promotion.date
                
                # If promotion has a step specified, use it directly
                if promotion.new_step is not None:
                    new_step = promotion.new_step
                else:
                    # Get current salary before promotion
                    current_salary = self.excel_handler.get_salary_value(
                        promotion_date, current_grade, current_step, self.sub_type
                    )
                    
                    # Calculate new step after promotion based on salary matching
                    new_step = self.excel_handler.find_equivalent_step(
                        current_salary, new_grade, promotion_date, self.sub_type
                    )
                
                # Get maximum step for new grade
                new_max_step = self.get_max_step_for_grade(new_grade)
                
                # Ensure step doesn't exceed maximum for new grade
                new_step = min(new_step, new_max_step)
                
                # Add to progression
                progression.append({
                    'date': promotion_date,
                    'grade': new_grade,
                    'step': new_step,
                    'event': promotion.promotion_type,
                    'salary': self.excel_handler.get_salary_value(
                        promotion_date, new_grade, new_step, self.sub_type
                    )
                })
                
                # Update current grade and step
                current_grade = new_grade
                current_step = new_step
                
                # Recalculate next increment date based on the promotion date
                next_increment_date = DateHandler.get_increment_date(promotion_date)
            
            # Apply annual increment
            if current_step < max_step:  # Only if not at max step for current grade
                new_step = current_step + 1
                
                # Add to progression
                progression.append({
                    'date': next_increment_date,
                    'grade': current_grade,
                    'step': new_step,
                    'event': 'Annual Increment',
                    'salary': self.excel_handler.get_salary_value(
                        next_increment_date, current_grade, new_step, self.sub_type
                    )
                })
                
                # Update current step
                current_step = new_step
                
                # Update last increment date
                last_increment_date = next_increment_date
            else:
                # If at max step, maintain the same step and add a record
                progression.append({
                    'date': next_increment_date,
                    'grade': current_grade,
                    'step': current_step,
                    'event': 'Maximum Step Maintained',
                    'salary': self.excel_handler.get_salary_value(
                        next_increment_date, current_grade, current_step, self.sub_type
                    )
                })
                # Update last increment date
                last_increment_date = next_increment_date
            
            # Process promotions that occur ON this increment date
            # (increment happens first, then promotion)
            promotions_on_increment = [p for p in self.promotions if p.date == next_increment_date]
            
            for promotion in sorted(promotions_on_increment, key=lambda p: p.date):
                # Apply promotion
                new_grade = promotion.new_grade
                promotion_date = promotion.date
                
                # If promotion has a step specified, use it directly
                if promotion.new_step is not None:
                    new_step = promotion.new_step
                else:
                    # Get current salary after increment but before promotion
                    current_salary = self.excel_handler.get_salary_value(
                        promotion_date, current_grade, current_step, self.sub_type
                    )
                    
                    # Calculate new step after promotion based on salary matching
                    new_step = self.excel_handler.find_equivalent_step(
                        current_salary, new_grade, promotion_date, self.sub_type
                    )
                
                # Get maximum step for new grade
                new_max_step = self.get_max_step_for_grade(new_grade)
                
                # Ensure step doesn't exceed maximum for new grade
                new_step = min(new_step, new_max_step)
                
                # Add to progression
                progression.append({
                    'date': promotion_date,
                    'grade': new_grade,
                    'step': new_step,
                    'event': promotion.promotion_type,
                    'salary': self.excel_handler.get_salary_value(
                        promotion_date, new_grade, new_step, self.sub_type
                    )
                })
                
                # Update current grade and step
                current_grade = new_grade
                current_step = new_step
                
                # Recalculate next increment date based on the promotion date
                next_increment_date = DateHandler.get_increment_date(promotion_date)
            
            # If no promotions changed the increment date, calculate the next one based on current date
            if next_increment_date <= progression[-1]['date']:
                if next_increment_date.month <= 6:
                    next_increment_date = datetime(next_increment_date.year + 1, 1, 1)
                else:
                    next_increment_date = datetime(next_increment_date.year + 1, 7, 1)
        
        return progression
    
    def _calculate_new_step_after_increment(self, grade, current_step, date):
        """Calculate the new step after an annual increment"""
        # Check if already at max step (15)
        if current_step >= 15:
            return current_step
            
        # For all units, annual increments add one step
        return current_step + 1
    
    def _calculate_new_step_after_promotion(self, old_grade, old_step, new_grade, date, promotion_type):
        """
        Calculate the new step after a promotion based on:
        1. Current salary value
        2. Unit-specific rules for promotion steps
        """
        # Get current salary
        current_salary = self.excel_handler.get_salary_value(
            date, old_grade, old_step, self.sub_type
        )
        
        if current_salary is None:
            # Default to step 1 if salary can't be determined
            return 1
        
        # Find equivalent step in new grade
        base_step = self.excel_handler.find_equivalent_step(
            current_salary, new_grade, date, self.sub_type
        )
        
        # Apply additional step based on unit rules
        additional_step = 0
        if self._should_add_promotion_step(self.unit_type, promotion_type):
            additional_step = 1
            
        new_step = base_step + additional_step
        
        # Ensure step doesn't exceed maximum (15)
        return min(new_step, 15)
    
    def _should_add_promotion_step(self, unit_type, promotion_type):
        """
        Determine if additional step should be added based on unit rules:
        - Mainstream/Tescom: Always add step at promotion/advancement
        - Subeb/Local Government: Add step only for actual promotions
        """
        if unit_type in ["Mainstream", "Tescom"]:
            # Always add step for Mainstream and Tescom
            return True
        else:  # Subeb or Local Government
            # Add step only if it's a promotion (not advancement/conversion)
            return promotion_type == "Promotion"
    
    def get_final_grade_and_step(self):
        """Get the final grade and step as of March 31, 2007"""
        progression = self.calculate_progression()
        if not progression:
            return None, None
            
        # Return the grade and step from the last record
        final_record = progression[-1]
        return final_record['grade'], final_record['step']


class SalaryProgressionApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Salary Progression Calculator")
        self.root.geometry("1000x800")
        
        self.excel_handler = ExcelDataHandler()
        self.calculator = SalaryProgressionCalculator(self.excel_handler)
        self.promotion_list = []
        
        # Add session management variables
        self.current_session = []
        self.session_exported = True  # Start with no active session
        self.export_file_path = None
        self.session_active = False  # New flag to track if session is active
        self.prompt_shown = False  # Flag to track if prompt has been shown
        
        # Add subtype cycling variables
        self.last_hatiss_hss = "HATISS"  # Track last selected between HATISS and HSS
        
        # Auto-load the Excel file from the directory
        self._auto_load_excel()
        
        self._setup_ui()
        
        # Show initial session prompt
        self._show_session_prompt()

    def _auto_load_excel(self):
        """Automatically load the Excel file from the directory"""
        try:
            # Get the directory of the current script
            script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            # Construct the path to the Excel file
            excel_path = os.path.join(script_dir, "salary_tables.xlsx")
            
            # Load the Excel file
            self.excel_handler.load_excel_file(excel_path)
            self.calculator = SalaryProgressionCalculator(self.excel_handler)
            
            # No need to update status as the UI is not yet set up
        except Exception as e:
            print(f"Failed to auto-load Excel file: {str(e)}")
            # We'll show an error message box after the UI is set up
            self.excel_load_error = str(e)    
            
    def _check_excel_load_status(self):
        """Check if there was an error loading the Excel file and display it"""
        if hasattr(self, 'excel_load_error'):
            messagebox.showerror("Error", f"Failed to load Excel file: {self.excel_load_error}")
            self.status_var.set("Error loading Excel file. Please check the file and restart the application.")
        else:
            self.status_var.set("Ready. Salary data loaded successfully.")        
    
    def _setup_ui(self):
        """Set up the main UI components"""
        # Create main container with scrollbar
        main_container = ttk.Frame(self.root)
        main_container.pack(fill=tk.BOTH, expand=True)

        # Create canvas and scrollbar
        canvas = tk.Canvas(main_container)
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        # Configure canvas
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack canvas and scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Add mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Create main frame inside scrollable area
        self.main_frame = ttk.Frame(scrollable_frame, padding=10)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Add close button to main window
        close_button = ttk.Button(self.main_frame, text="Close Application", command=self._close_application)
        close_button.pack(side=tk.TOP, anchor=tk.E, padx=5, pady=5)

        # Create validation commands early, before they're used
        validate_date = self.root.register(self._validate_date_input)
        validate_alpha = self.root.register(self._validate_alpha_input)
        validate_numeric = self.root.register(self._validate_numeric_input)
        
        # Personal Information frame
        personal_frame = ttk.LabelFrame(self.main_frame, text="Personal Information", padding=5)
        personal_frame.pack(fill=tk.X, pady=5)

        # Name field - alphabetic characters only
        ttk.Label(personal_frame, text="Name:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.name_var = tk.StringVar()
        self.name_entry = ttk.Entry(personal_frame, textvariable=self.name_var, width=30, 
                validate="key", validatecommand=(validate_alpha, '%P'))
        self.name_entry.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        self.name_entry.bind('<Key>', lambda e: self._validate_personal_info())
        self.name_entry.bind('<Button-1>', lambda e: self._validate_personal_info())
        
        # Set initial focus to name entry field
        self.name_entry.focus_set()

        # Oracle Number - numeric characters only
        ttk.Label(personal_frame, text="Oracle Number:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        self.oracle_number_var = tk.StringVar()
        self.oracle_entry = ttk.Entry(personal_frame, textvariable=self.oracle_number_var, width=20, 
                validate="key", validatecommand=(validate_numeric, '%P'))
        self.oracle_entry.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        self.oracle_entry.bind('<Key>', lambda e: self._validate_personal_info())
        self.oracle_entry.bind('<Button-1>', lambda e: self._validate_personal_info())

        # Sex dropdown
        ttk.Label(personal_frame, text="Sex:").grid(row=0, column=4, sticky=tk.W, padx=5, pady=5)
        self.sex_var = tk.StringVar()
        sex_options = ["M", "F"]
        self.sex_dropdown = ttk.Combobox(personal_frame, textvariable=self.sex_var, values=sex_options, 
                            state="readonly", width=5)
        self.sex_dropdown.grid(row=0, column=5, sticky=tk.W, padx=5, pady=5)
        self.sex_dropdown.bind('<Key>', lambda e: self._validate_personal_info())
        self.sex_dropdown.bind('<Button-1>', lambda e: self._validate_personal_info())
        
        # Bind keyboard shortcuts for sex selection
        def handle_sex_key(event):
            if event.char.upper() == 'M':
                self.sex_var.set("M")
            elif event.char.upper() == 'F':
                self.sex_var.set("F")
        self.sex_dropdown.bind('<Key>', handle_sex_key)

        # Date of Birth
        ttk.Label(personal_frame, text="Date of Birth:").grid(row=0, column=6, sticky=tk.W, padx=5, pady=5)
        self.dob_var = tk.StringVar()
        self.dob_entry = ttk.Entry(personal_frame, textvariable=self.dob_var, width=15, 
                validate="key", validatecommand=(validate_date, '%P'))
        self.dob_entry.grid(row=0, column=8, sticky=tk.W, padx=5, pady=5)
        self.dob_entry.bind('<Key>', lambda e: self._validate_personal_info())
        self.dob_entry.bind('<Button-1>', lambda e: self._validate_personal_info())
        ttk.Label(personal_frame, text="(DD-MM-YY)").grid(row=0, column=7, sticky=tk.W, padx=0, pady=5)
        
        # Basic information frame
        info_frame = ttk.LabelFrame(self.main_frame, text="Employee Information", padding=5)
        info_frame.pack(fill=tk.X, pady=5)
        
        # Unit selection
        ttk.Label(info_frame, text="Unit:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.unit_var = tk.StringVar()
        unit_options = ["Mainstream", "Local Government", "Tescom", "Subeb"]
        self.unit_dropdown = ttk.Combobox(info_frame, textvariable=self.unit_var, values=unit_options, state="readonly", width=20)
        self.unit_dropdown.grid(row=0, column=1, sticky=tk.W, padx=5, pady=5)
        self.unit_dropdown.bind("<<ComboboxSelected>>", self._on_unit_change)
        self.unit_dropdown.bind('<FocusIn>', lambda e: self._validate_personal_info())
        
        # Bind keyboard shortcuts for unit selection
        def handle_unit_key(event):
            key = event.char.upper()
            if key == 'M':
                self.unit_var.set("Mainstream")
                self._on_unit_change(None)
            elif key == 'L':
                self.unit_var.set("Local Government")
                self._on_unit_change(None)
            elif key == 'S':
                self.unit_var.set("Subeb")
                self._on_unit_change(None)
            elif key == 'T':
                self.unit_var.set("Tescom")
                self._on_unit_change(None)
        self.unit_dropdown.bind('<Key>', handle_unit_key)
        
        # Sub-type selection (for Mainstream and Tescom)
        ttk.Label(info_frame, text="Sub-Type:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        self.subtype_var = tk.StringVar()
        self.subtype_options = ["Standard", "HATISS", "HSS", "MSS"]
        self.subtype_dropdown = ttk.Combobox(info_frame, textvariable=self.subtype_var, 
                                          values=self.subtype_options, state="disabled", width=20)
        self.subtype_dropdown.grid(row=0, column=3, sticky=tk.W, padx=5, pady=5)
        
        # Bind keyboard shortcuts for subtype selection
        def handle_subtype_key(event):
            if not self.subtype_dropdown['state'] == 'disabled':  # Only process if enabled
                key = event.char.upper()
                if key == 'S':
                    self.subtype_var.set("Standard")
                elif key == 'H':
                    # Toggle between HATISS and HSS
                    if self.last_hatiss_hss == "HATISS":
                        self.subtype_var.set("HSS")
                        self.last_hatiss_hss = "HSS"
                    else:
                        self.subtype_var.set("HATISS")
                        self.last_hatiss_hss = "HATISS"
                elif key == 'M':
                    self.subtype_var.set("MSS")
        self.subtype_dropdown.bind('<Key>', handle_subtype_key)
        
        # Appointment date
        ttk.Label(info_frame, text="Date of Appointment:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.appointment_date_var = tk.StringVar()
        
        # Create a validation command
        validate_date = self.root.register(self._validate_date_input)
        
        # Apply validation to the appointment date entry field
        ttk.Entry(info_frame, textvariable=self.appointment_date_var, width=15, 
                validate="key", validatecommand=(validate_date, '%P')).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
            
        # Initial grade and step
        ttk.Label(info_frame, text="Initial Grade Level:").grid(row=1, column=2, sticky=tk.W, padx=5, pady=5)
        self.initial_grade_var = tk.IntVar(value=1)
        grade_options = list(range(1, 18))
        self.initial_grade_combo = ttk.Combobox(info_frame, textvariable=self.initial_grade_var, values=grade_options, 
                        width=5)
        self.initial_grade_combo.grid(row=1, column=3, sticky=tk.W, padx=5, pady=5)
        
        # Add validation for grade input
        validate_grade = self.root.register(self._validate_grade_input)
        self.initial_grade_combo.configure(validate="key", validatecommand=(validate_grade, '%P'))
        
        ttk.Label(info_frame, text="Initial Step:").grid(row=1, column=4, sticky=tk.W, padx=5, pady=5)
        self.initial_step_var = tk.IntVar(value=1)
        step_options = list(range(1, 16))
        initial_step_combo = ttk.Combobox(info_frame, textvariable=self.initial_step_var, values=step_options, 
                        width=5)
        initial_step_combo.grid(row=1, column=5, sticky=tk.W, padx=5, pady=5)
        
        # Container to hold both promotion history and progression results side by side
        history_results_frame = ttk.Frame(self.main_frame)
        history_results_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Promotion History Section with improved layout
        promotion_frame = ttk.LabelFrame(history_results_frame, text="Promotion/Advancement History", padding=5)
        promotion_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Create a container frame for the promotion section
        promotion_container = ttk.Frame(promotion_frame)
        promotion_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Promotion entry fields
        entry_frame = ttk.Frame(promotion_container)
        entry_frame.pack(fill=tk.X, pady=5)

        # Date field
        ttk.Label(entry_frame, text="Date:").grid(row=0, column=0, sticky=tk.W, padx=5)
        self.promotion_date_var = tk.StringVar()
        self.promotion_date_entry = ttk.Entry(entry_frame, textvariable=self.promotion_date_var, width=15,
                                            validate="key", validatecommand=(validate_date, '%P'))
        self.promotion_date_entry.grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Label(entry_frame, text="(DD-MM-YY)").grid(row=0, column=2, sticky=tk.W, padx=0)
        
        # Add event handlers for promotion fields
        self.promotion_date_entry.bind('<FocusIn>', self._prompt_new_session)
        self.promotion_date_entry.bind('<Button-1>', self._prompt_new_session)
        self.promotion_date_entry.bind('<Return>', lambda e: self._handle_promotion_add())

        # New Grade field
        ttk.Label(entry_frame, text="New Grade:").grid(row=0, column=3, sticky=tk.W, padx=5)
        self.promotion_grade_var = tk.IntVar()
        self.promotion_grade_combo = ttk.Combobox(entry_frame, textvariable=self.promotion_grade_var, 
                                            values=grade_options, width=5)
        self.promotion_grade_combo.grid(row=0, column=4, sticky=tk.W, padx=5)
        
        # Add validation for grade input
        self.promotion_grade_combo.configure(validate="key", validatecommand=(validate_grade, '%P'))
        
        # Add event handlers for grade combo
        self.promotion_grade_combo.bind('<FocusIn>', self._prompt_new_session)
        self.promotion_grade_combo.bind('<Button-1>', self._prompt_new_session)
        self.promotion_grade_combo.bind('<Return>', lambda e: self._handle_promotion_add())

        # New Step field (optional)
        ttk.Label(entry_frame, text="New Step:").grid(row=0, column=5, sticky=tk.W, padx=5)
        self.promotion_step_var = tk.StringVar()  # Change from IntVar to StringVar
        self.promotion_step_combo = ttk.Combobox(entry_frame, textvariable=self.promotion_step_var, 
                                            values=step_options, width=5)
        self.promotion_step_combo.grid(row=0, column=6, sticky=tk.W, padx=5)
        
        # Add event handlers for step combo
        self.promotion_step_combo.bind('<FocusIn>', self._prompt_new_session)
        self.promotion_step_combo.bind('<Button-1>', self._prompt_new_session)
        self.promotion_step_combo.bind('<Return>', lambda e: self._handle_promotion_add())

        # Type field
        ttk.Label(entry_frame, text="Type:").grid(row=0, column=7, sticky=tk.W, padx=5)
        self.promotion_type_var = tk.StringVar(value="Promotion")
        self.promotion_type_combo = ttk.Combobox(entry_frame, textvariable=self.promotion_type_var, 
                                            values=["Promotion", "Advancement", "Conversion", "Harmonization"], 
                                            state="readonly", width=15)
        self.promotion_type_combo.grid(row=0, column=8, sticky=tk.W, padx=5)
        
        # Add keyboard shortcuts for type selection
        def handle_type_key(event):
            key = event.char.upper()
            if key == 'P':
                self.promotion_type_var.set("Promotion")
            elif key == 'A':
                self.promotion_type_var.set("Advancement")
            elif key == 'C':
                self.promotion_type_var.set("Conversion")
            elif key == 'H':
                self.promotion_type_var.set("Harmonization")
            return "break"  # Prevent the key from being processed further
        self.promotion_type_combo.bind('<Key>', handle_type_key)
        
        # Add event handlers for type combo
        self.promotion_type_combo.bind('<FocusIn>', self._prompt_new_session)
        self.promotion_type_combo.bind('<Button-1>', self._prompt_new_session)
        self.promotion_type_combo.bind('<Return>', lambda e: self._handle_promotion_add())

        # Add Promotion Button
        self.add_promotion_button = ttk.Button(entry_frame, text="Add Promotion", command=self._handle_promotion_add)
        self.add_promotion_button.grid(row=0, column=9, sticky=tk.W, padx=5)
        
        # Add event handlers for add button
        self.add_promotion_button.bind('<Button-1>', self._prompt_new_session)
        self.add_promotion_button.bind('<Return>', lambda e: self._handle_promotion_add())

        # Create a frame for the promotion treeview
        tree_frame = ttk.Frame(promotion_container)
        tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)

        # Promotion history treeview
        columns = ('date', 'grade', 'step', 'type')
        self.promotion_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=10)
        self.promotion_tree.heading('date', text='Date')
        self.promotion_tree.heading('grade', text='New Grade')
        self.promotion_tree.heading('step', text='New Step')
        self.promotion_tree.heading('type', text='Type')

        # Adjust column widths for better visibility
        self.promotion_tree.column('date', width=100)
        self.promotion_tree.column('grade', width=100)
        self.promotion_tree.column('step', width=100)
        self.promotion_tree.column('type', width=150)

        self.promotion_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.promotion_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.promotion_tree.configure(yscrollcommand=scrollbar.set)

        # Add buttons for managing promotions
        button_frame = ttk.Frame(promotion_container)
        button_frame.pack(fill=tk.X, pady=5)

        ttk.Button(button_frame, text="Remove Selected", command=self._remove_promotion).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Edit Selected", command=self._edit_promotion).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Clear All", command=self._clear_promotions).pack(side=tk.LEFT, padx=5)

        # Progression Results Section (Now side-by-side with Promotion History)
        results_frame = ttk.LabelFrame(history_results_frame, text="Progression Results", padding=5)
        results_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Add show/hide button frame
        results_button_frame = ttk.Frame(results_frame)
        results_button_frame.pack(fill=tk.X, pady=5)

        # Create show/hide button
        self.show_results_button = ttk.Button(results_button_frame, text="Show Results", command=self._toggle_results)
        self.show_results_button.pack(side=tk.LEFT, padx=5)

        # Remove the container frame since we'll use a popup window instead
        self.results_popup = None
        
        # Calculate button and final status display
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(fill=tk.X, pady=10)

        # Create a container frame to control layout
        final_status_container = ttk.Frame(button_frame)
        final_status_container.pack(fill=tk.X, pady=5)

        # Buttons section (left side)
        button_section = ttk.Frame(final_status_container)
        button_section.pack(side=tk.LEFT, padx=5)

        ttk.Button(button_section, text="Calculate Progression", command=self._calculate_progression).pack(side=tk.TOP, pady=2)
        ttk.Button(button_section, text="Clear All", command=self._clear_all).pack(side=tk.TOP, pady=2)

        # Final Status Display with fixed width
        final_status_frame = ttk.LabelFrame(final_status_container, text="Final Status as of 31st March, 2007", padding=5)
        final_status_frame.pack(side=tk.LEFT, fill=tk.BOTH, padx=10, pady=5)
        
        # Create a container frame for the status label
        status_container = ttk.Frame(final_status_frame)
        status_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Use a larger font for better visibility
        final_status_font = ('Arial', 12, 'bold')

        self.final_status_var = tk.StringVar(value="")
        self.final_status_label = ttk.Label(
            status_container, 
            textvariable=self.final_status_var,
            font=final_status_font,
            background="#f0f0f0",  # Light gray background
            relief=tk.SUNKEN,
            padding=10,
            anchor=tk.CENTER,
            width=40  # Adjusted width for better visibility
        )
        self.final_status_label.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add Save for Export button in a new frame below the final status
        save_button_frame = ttk.Frame(final_status_container)
        save_button_frame.pack(fill=tk.X, pady=5)
        
        save_button = ttk.Button(save_button_frame, text="Save for Export", command=self._save_for_export)
        save_button.pack(side=tk.RIGHT, padx=5, pady=5)

        # Add spacer frame to make layout consistent with other sections
        spacer_frame = ttk.Frame(final_status_container)
        spacer_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # Status bar
        self.status_var = tk.StringVar()
        status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Check Excel load status
        self.root.after(100, self._check_excel_load_status)
        
        # Add Export Section
        export_frame = ttk.LabelFrame(self.main_frame, text="Export Management", padding=5)
        export_frame.pack(fill=tk.X, pady=5)
        
        # Export buttons
        button_frame = ttk.Frame(export_frame)
        button_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(button_frame, text="New Session", command=self._new_session).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Continue Session", command=self._continue_session).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="End Session", command=self._end_session).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Export Session", command=self._export_session).pack(side=tk.LEFT, padx=5)
        
        # Session status label
        self.session_status_var = tk.StringVar(value="No active session")
        ttk.Label(export_frame, textvariable=self.session_status_var).pack(pady=5)

        # Set minimum window size
        self.root.minsize(1000, 800)
    
    def _validate_date_input(self, value):
        """Validate date input to accept only numbers and hyphens"""
        # Allow empty strings (for field clearing)
        if value == "":
            return True
        
        # Check if the input string contains only digits and hyphens
        return all(c.isdigit() or c == '-' for c in value)

    def _validate_session(self):
        """Validate if a session is active before allowing any actions"""
        if self.session_active:
            return True

        if not self.prompt_shown:
            self.prompt_shown = True
            messagebox.showinfo("Session Required", 
                            "Please start a new session before making any entries.\n\n"
                            "Click 'New Session' to begin.")
        return False

    def _validate_personal_info(self):
        """Validate personal information fields access"""
        if not self.session_active:
            messagebox.showinfo("Session Required", 
                            "Please start a new session before making any entries.\n\n"
                            "Click 'New Session' to begin.")
            return False
        return True

    def _handle_promotion_add(self, event=None):
        """Handle promotion addition with proper session validation"""
        if not self.session_active:
            self._prompt_new_session()
            return "break"
        self._add_promotion()
        return "break"

    def _add_promotion(self):
        """Add a promotion entry"""
        if not self._validate_session():
            return
            
        # Make sure the button is in "Add" mode
        self.add_promotion_button.config(text="Add Promotion", command=self._add_promotion)
        
        try:
            # Validate inputs
            date_str = self.promotion_date_var.get()
            promotion_type = self.promotion_type_var.get()
            new_grade = self.promotion_grade_var.get()
            
            # Handle step value carefully
            step_value = self.promotion_step_var.get().strip()
            new_step = None
            if step_value:
                try:
                    new_step = int(step_value)
                except (ValueError, TypeError):
                    new_step = None
            
            if not date_str or not promotion_type or not new_grade:
                messagebox.showerror("Error", "Please fill all required promotion details")
                return
                
            try:
                promotion_date = DateHandler.parse_date(date_str)
                if not promotion_date:
                    raise ValueError("Invalid date format")
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid date: {str(e)}")
                return
            
            # Validate promotion date is between appointment and end date
            appointment_date_str = self.appointment_date_var.get()
            if appointment_date_str:
                try:
                    appointment_date = DateHandler.parse_date(appointment_date_str)
                    if promotion_date < appointment_date:
                        messagebox.showerror("Error", "Promotion date cannot be before appointment date")
                        return
                    if promotion_date > datetime(2007, 3, 31):
                        messagebox.showerror("Error", "Promotion date cannot be after March 31, 2007")
                        return
                except ValueError:
                    pass  # Will be caught during calculation
            
            # Create promotion entry with optional step
            promotion = PromotionEntry(promotion_date, promotion_type, new_grade, new_step)
            
            # Add to treeview
            item_id = self.promotion_tree.insert('', 'end', values=(
                DateHandler.format_date(promotion_date),
                new_grade,
                new_step if new_step is not None else '',  # Show empty string if no step specified
                promotion_type
            ))
            
            # Store promotion with its treeview ID
            self.promotion_list.append({
                'id': item_id,
                'promotion': promotion
            })
            
            # Clear entry fields
            self.promotion_date_var.set('')
            self.promotion_grade_var.set('')
            self.promotion_step_var.set('')  # Clear step field
            
            self.status_var.set(f"Added {promotion_type} to Grade {new_grade}" + 
                             (f" Step {new_step}" if new_step is not None else ""))
            
            # Set focus back to the date entry field
            self.promotion_date_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to add promotion: {str(e)}")
    
    def _browse_excel(self):
        """Open file dialog to select Excel file"""
        file_path = filedialog.askopenfilename(
            title="Select Salary Table Excel File",
            filetypes=[("Excel files", "*.xlsx;*.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path_var.set(file_path)
    
    def _load_excel_data(self):
        """Load the selected Excel file"""
        file_path = self.excel_path_var.get()
        if not file_path:
            messagebox.showerror("Error", "Please select an Excel file first")
            return
            
        try:
            self.status_var.set("Loading Excel data...")
            self.root.update()
            
            self.excel_handler.load_excel_file(file_path)
            self.calculator = SalaryProgressionCalculator(self.excel_handler)
            
            self.status_var.set(f"Successfully loaded salary data from {os.path.basename(file_path)}")
            messagebox.showinfo("Success", "Salary data loaded successfully")
        except Exception as e:
            self.status_var.set("Error loading Excel file")
            messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")
    
    def _on_unit_change(self, event):
        """Handle unit selection change"""
        unit = self.unit_var.get()
        if unit in ["Mainstream", "Tescom"]:
            self.subtype_dropdown.configure(state="readonly")
            # Reset subtype to Standard when unit changes
            self.subtype_var.set("Standard")
            self.last_hatiss_hss = "HATISS"  # Reset HATISS/HSS toggle
        else:
            self.subtype_var.set("")
            self.subtype_dropdown.configure(state="disabled")
    
        # Update grade options based on unit and sub-type
        self._update_grade_options()
    
    def _on_subtype_change(self, event):
        """Handle sub-type selection change"""
        self._update_grade_options()
    
    def _update_grade_options(self):
        """Update grade options based on unit and sub-type"""
        unit = self.unit_var.get()
        sub_type = self.subtype_var.get()
        
        # Default grade options (1-17)
        grade_options = list(range(1, 18))
        
        # Remove grade 11 for Subeb and Local Government
        if unit in ["Subeb", "Local Government"]:
            grade_options.remove(11)
        # Remove grade 11 for Mainstream and Tescom with Standard sub-type
        elif unit in ["Mainstream", "Tescom"] and sub_type == "Standard":
            grade_options.remove(11)
        
        # Update grade combo boxes
        self.initial_grade_combo['values'] = grade_options
        self.promotion_grade_combo['values'] = grade_options
        
        # Safely handle current grade values
        try:
            current_initial_grade = self.initial_grade_var.get()
            if current_initial_grade == 11 and 11 not in grade_options:
                self.initial_grade_var.set(1)
        except (ValueError, tk.TclError):
            self.initial_grade_var.set(1)
            
        try:
            current_promotion_grade = self.promotion_grade_var.get()
            if current_promotion_grade == 11 and 11 not in grade_options:
                self.promotion_grade_var.set(1)
        except (ValueError, tk.TclError):
            self.promotion_grade_var.set(1)
    
    def _validate_grade_input(self, value):
        """Validate grade input to prevent grade 11 in certain cases"""
        if not value:
            return True
            
        try:
            grade = int(value)
            unit = self.unit_var.get()
            sub_type = self.subtype_var.get()
            
            # Check if grade 11 is allowed
            if grade == 11:
                if unit in ["Subeb", "Local Government"]:
                    messagebox.showerror("Error", "Grade Level 11 does not exist on current Salary scale")
                    return False
                elif unit in ["Mainstream", "Tescom"] and sub_type == "Standard":
                    messagebox.showerror("Error", "Grade Level 11 does not exist on current Salary scale")
                    return False
            
            return True
        except ValueError:
            return False
    
    def _remove_promotion(self):
        """Remove the selected promotion entry"""
        selected_item = self.promotion_tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select a promotion to remove")
            return
            
        item_id = selected_item[0]
        
        # Find the promotion in our list
        for i, item in enumerate(self.promotion_list):
            if item['id'] == item_id:
                self.promotion_list.pop(i)
                self.promotion_tree.delete(item_id)
                self.status_var.set("Removed promotion entry")
                
                # Reset the Add Promotion button back to Add mode if it's in Update mode
                if self.add_promotion_button['text'] == "Update Promotion":
                    self.add_promotion_button.config(text="Add Promotion", command=self._add_promotion)
                    self.promotion_date_var.set('')
                    self.promotion_grade_var.set('')
                return
            
    def _edit_promotion(self):
        """Edit the selected promotion entry"""
        selected_item = self.promotion_tree.selection()
        if not selected_item:
            messagebox.showinfo("Info", "Please select a promotion to edit")
            return
            
        item_id = selected_item[0]
        
        # Find the promotion in our list
        for i, item in enumerate(self.promotion_list):
            if item['id'] == item_id:
                # Get the current values
                promotion = item['promotion']
                
                # Fill the form with current values
                self.promotion_date_var.set(DateHandler.format_date(promotion.date))
                self.promotion_type_var.set(promotion.promotion_type)
                self.promotion_grade_var.set(promotion.new_grade)
                
                # Change the Add Promotion button to Update Promotion
                self.add_promotion_button.config(text="Update Promotion", command=lambda idx=i, item_id=item_id: self._update_promotion(idx, item_id))
                
                # Set focus to the date field
                self.promotion_date_entry.focus_set()
                return
    def _update_promotion(self, idx, item_id):
        """Update an existing promotion with new values"""
        try:
            # Validate inputs
            date_str = self.promotion_date_var.get()
            promotion_type = self.promotion_type_var.get()
            new_grade = self.promotion_grade_var.get()
            
            # Handle step value carefully
            step_value = self.promotion_step_var.get()
            new_step = None
            if step_value and step_value != '':
                try:
                    new_step = int(step_value)
                except (ValueError, TypeError):
                    new_step = None
            
            if not date_str or not promotion_type or not new_grade:
                messagebox.showerror("Error", "Please fill all required promotion details")
                return
                
            try:
                promotion_date = DateHandler.parse_date(date_str)
                if not promotion_date:
                    raise ValueError("Invalid date format")
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid date: {str(e)}")
                return
            
            # Validate promotion date is between appointment and end date
            appointment_date_str = self.appointment_date_var.get()
            if appointment_date_str:
                try:
                    appointment_date = DateHandler.parse_date(appointment_date_str)
                    if promotion_date < appointment_date:
                        messagebox.showerror("Error", "Promotion date cannot be before appointment date")
                        return
                    if promotion_date > datetime(2007, 3, 31):
                        messagebox.showerror("Error", "Promotion date cannot be after March 31, 2007")
                        return
                except ValueError:
                    pass  # Will be caught during calculation
            
            # Create updated promotion entry with optional step
            updated_promotion = PromotionEntry(promotion_date, promotion_type, new_grade, new_step)
            
            # Update the promotion in our list
            self.promotion_list[idx]['promotion'] = updated_promotion
            
            # Update the treeview
            self.promotion_tree.item(item_id, values=(
                DateHandler.format_date(promotion_date),
                new_grade,
                new_step if new_step is not None else '',  # Show empty string if no step specified
                promotion_type
            ))
            
            # Reset form and button
            self.promotion_date_var.set('')
            self.promotion_grade_var.set('')
            self.promotion_step_var.set('')  # Clear step field
            self.add_promotion_button.config(text="Add Promotion", command=self._add_promotion)
            
            self.status_var.set(f"Updated {promotion_type} to Grade {new_grade}" + 
                             (f" Step {new_step}" if new_step is not None else ""))
            
            # Set focus back to the date entry field
            self.promotion_date_entry.focus_set()
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update promotion: {str(e)}")  
        
    def _clear_all(self):
        """Clear all inputs and results"""
        # Clear basic information
        self.unit_var.set('')
        self.subtype_var.set('')
        self.appointment_date_var.set('')
        self.initial_grade_var.set(1)
        self.initial_step_var.set(1)
        
        # Clear personal information
        self.name_var.set('')
        self.oracle_number_var.set('')
        self.sex_var.set('')
        self.dob_var.set('')
        
        # Clear promotion entries
        for item in self.promotion_tree.get_children():
            self.promotion_tree.delete(item)
        self.promotion_list = []

        # Reset the button to "Add" mode
        self.add_promotion_button.config(text="Add Promotion", command=self._add_promotion)
        
        # Clear final status display
        self.final_status_var.set("")
        
        self.status_var.set("All inputs cleared")
    
    def _calculate_progression(self):
        """Calculate and display the salary progression"""
        if not self._validate_session():
            return
            
        try:
            # Validate basic inputs
            unit = self.unit_var.get()
            sub_type = self.subtype_var.get() if unit in ["Mainstream", "Tescom"] else None
            appointment_date_str = self.appointment_date_var.get()
            initial_grade = self.initial_grade_var.get()
            initial_step = self.initial_step_var.get()
            
            if not unit:
                messagebox.showerror("Error", "Please select a unit")
                return
                
            if unit in ["Mainstream", "Tescom"] and not sub_type:
                messagebox.showerror("Error", "Please select a sub-type for this unit")
                return
                
            if not appointment_date_str:
                messagebox.showerror("Error", "Please enter date of appointment")
                return
            
            try:
                appointment_date = DateHandler.parse_date(appointment_date_str)
                if not appointment_date:
                    raise ValueError("Could not parse appointment date")
                if appointment_date > datetime(2007, 3, 31):
                    messagebox.showerror("Error", "Appointment date cannot be after March 31, 2007")
                    return
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid appointment date: {str(e)}")
                return
            
            # Set calculator parameters
            self.calculator.set_parameters(
                unit, sub_type, appointment_date, initial_grade, initial_step
            )
            
            # Add promotions to calculator
            self.calculator.promotions = []
            for item in self.promotion_list:
                self.calculator.add_promotion(item['promotion'])
            
            # Calculate progression
            self.status_var.set("Calculating salary progression...")
            self.root.update()
            
            progression = self.calculator.calculate_progression()
            
            # Store progression data for popup window
            self.progression_data = progression
            
            # Get final grade and step
            final_grade, final_step = self.calculator.get_final_grade_and_step()
            
            # Format the final status in the required format
            final_status = f"Grade Level {final_grade} Step {final_step}"
            
            # Update status bar with calculation completion message
            self.status_var.set(f"Calculation complete.")
            
            # Set the final status display
            self.final_status_var.set(final_status)
            
            # Highlight the final status display
            self.final_status_label.configure(background="#e6ffe6")  # Light green background
            
            # After 1 second, return to normal background
            self.root.after(1000, lambda: self.final_status_label.configure(background="#f0f0f0"))
            
            # After successful calculation, add to session if one is active
            if not self.session_exported:
                self._add_to_session()
            
        except Exception as e:
            self.status_var.set("Error during calculation")
            self.final_status_var.set("Error occurred during calculation")
            messagebox.showerror("Calculation Error", str(e))
            import traceback
            traceback.print_exc()

    def _new_session(self):
        """Start a new export session"""
        if not self.session_exported and self.current_session:
            if not messagebox.askyesno("Warning", "Previous session not exported, continue anyway?"):
                return
        
        self.current_session = []
        self.session_exported = False
        self.session_active = True
        self.prompt_shown = False  # Reset prompt flag
        self.session_status_var.set("New session started")
        
        # Enable all input fields
        self._toggle_input_fields(True)
        
        # Clear personal information fields
        self.name_var.set('')
        self.oracle_number_var.set('')
        self.sex_var.set('')
        self.dob_var.set('')
        
        # Clear other fields
        self.unit_var.set('')
        self.subtype_var.set('')
        self.appointment_date_var.set('')
        self.initial_grade_var.set(1)
        self.initial_step_var.set(1)
        
        # Clear promotion entries
        for item in self.promotion_tree.get_children():
            self.promotion_tree.delete(item)
        self.promotion_list = []
        
        # Add current calculation to session if available
        if hasattr(self, 'final_status_var') and self.final_status_var.get():
            self._add_to_session()
            
        # Schedule focus set after a short delay to ensure window is updated
        self.root.after(100, lambda: self.name_entry.focus_set())

    def _continue_session(self):
        """Continue with the previous session"""
        self.session_status_var.set("Continuing previous session")
        if hasattr(self, 'final_status_var') and self.final_status_var.get():
            self._add_to_session()

    def _end_session(self):
        """End the current session"""
        if not self.current_session:
            messagebox.showinfo("Info", "No active session to end")
            return
            
        self.session_status_var.set("Session ended")
        self.session_exported = True
        self.session_active = False
        self.prompt_shown = False  # Reset prompt flag
        
        # Clear the session after ending
        self.current_session = []
        self.session_status_var.set("No active session")

    def _add_to_session(self):
        """Add current calculation to the session"""
        if not self.session_exported:
            # Extract data from current calculation
            session_data = {
                'oracle_number': self.oracle_number_var.get(),
                'sex': self.sex_var.get(),
                'name': self.name_var.get(),
                'dob': self.dob_var.get(),
                'appointment_date': self.appointment_date_var.get(),
                'final_status': self.final_status_var.get(),
                'unit': self.unit_var.get()  # Add unit information to session data
            }
            
            # Check if this entry already exists in the session
            if session_data not in self.current_session:
                self.current_session.append(session_data)
                self.session_status_var.set(f"Session contains {len(self.current_session)} entries")
            else:
                self.status_var.set("Entry already exists in session")

    def _load_grade_values(self):
        """Load grade values from Excel file"""
        try:
            # Use the full path to the grade values Excel file
            grade_values_path = r"D:\VISUAL STUDIO CODE\LASPEC SALARY PROGRESSION\grade_values.xlsx"
            
            if not os.path.exists(grade_values_path):
                raise FileNotFoundError("Grade values Excel file not found")
            
            # Load the Excel file
            grade_values_df = pd.read_excel(grade_values_path)
            
            # Create a dictionary to store grade and step values
            self.grade_values = {}
            
            # Process each row in the grade values sheet
            for _, row in grade_values_df.iterrows():
                grade_step = row['GRADE LEVEL AND STEP']
                basic_salary = row['BASIC SALARY']
                pensionable_allowance = row['PENSIONABLE ALLOWANCE']
                
                # Store the values in the dictionary
                self.grade_values[grade_step] = {
                    'basic_salary': basic_salary,
                    'pensionable_allowance': pensionable_allowance
                }
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load grade values: {str(e)}")
            self.grade_values = {}

    def _export_session(self):
        """Export the current session to Excel"""
        if not self.current_session:
            messagebox.showinfo("Info", "No data to export")
            return
            
        try:
            # Load grade values
            self._load_grade_values()
            
            # Create a new Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            
            # Add headers with new columns
            headers = ['S/N', 'AGENCY CODE', 'ORACLE NO', 'SEX', "PARTICIPANT'S NAME", 'DATE OF BIRTH', 
                      'DATE OF EMPLOYMENT', 'GRADE LEVEL AND STEP', 'BASIC SALARY', 'PENSIONABLE ALLOWANCE']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                # Left align headers
                cell.alignment = openpyxl.styles.Alignment(horizontal='left')
            
            # Add data
            for row, data in enumerate(self.current_session, 2):
                # Add S/N (row number)
                ws.cell(row=row, column=1, value=row-1)
                
                # Add Agency Code based on unit type
                unit = data.get('unit', '')
                agency_code = None
                if unit == 'Subeb':
                    agency_code = 1  # Set as integer for Subeb
                elif unit == 'Local Government':
                    agency_code = 'L'  # Keep as string for Local Government
                
                # Set cell value with appropriate format and alignment
                cell = ws.cell(row=row, column=2)
                if agency_code is not None:
                    if isinstance(agency_code, int):
                        cell.value = agency_code
                        cell.number_format = '0'  # Format as number
                    else:
                        cell.value = agency_code
                        cell.number_format = '@'  # Format as text
                    # Left align the cell
                    cell.alignment = openpyxl.styles.Alignment(horizontal='left')
                
                # Format Oracle No as number
                oracle_no = data['oracle_number']
                if oracle_no:
                    ws.cell(row=row, column=3, value=int(oracle_no))
                
                # Add other data
                ws.cell(row=row, column=4, value=data['sex'])
                ws.cell(row=row, column=5, value=data['name'])
                
                # Format dates
                try:
                    # Format Date of Birth
                    dob = DateHandler.parse_date(data['dob'])
                    if dob:
                        ws.cell(row=row, column=6, value=dob).number_format = 'd-mmm-yy'
                    
                    # Format Date of Employment
                    appointment_date = DateHandler.parse_date(data['appointment_date'])
                    if appointment_date:
                        ws.cell(row=row, column=7, value=appointment_date).number_format = 'd-mmm-yy'
                except ValueError:
                    # If date parsing fails, keep original format
                    ws.cell(row=row, column=6, value=data['dob'])
                    ws.cell(row=row, column=7, value=data['appointment_date'])
                
                # Format final status (Grade Level and Step)
                final_status = data['final_status']
                # Extract grade and step numbers
                grade_match = re.search(r'Grade Level (\d+)', final_status)
                step_match = re.search(r'Step (\d+)', final_status)
                if grade_match and step_match:
                    grade = grade_match.group(1)
                    step = step_match.group(1)
                    grade_step = f"{grade} {step}"
                    ws.cell(row=row, column=8, value=grade_step)
                    
                    # Get corresponding salary and allowance values
                    if grade_step in self.grade_values:
                        values = self.grade_values[grade_step]
                        # Format basic salary with accounting format (no currency symbol)
                        basic_salary_cell = ws.cell(row=row, column=9)
                        basic_salary_cell.value = values['basic_salary']
                        basic_salary_cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                        
                        # Format pensionable allowance with accounting format (no currency symbol)
                        allowance_cell = ws.cell(row=row, column=10)
                        allowance_cell.value = values['pensionable_allowance']
                        allowance_cell.number_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
                else:
                    ws.cell(row=row, column=8, value=final_status)
            
            # Adjust column widths based on content
            for col in range(1, len(headers) + 1):
                max_length = 0
                column = openpyxl.utils.get_column_letter(col)
                
                # Check header length
                header_length = len(str(headers[col-1]))
                max_length = max(max_length, header_length)
                
                # Check data length
                for row in range(2, len(self.current_session) + 2):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        cell_length = len(str(cell_value))
                        max_length = max(max_length, cell_length)
                
                # Set column width (adding some padding)
                ws.column_dimensions[column].width = max_length + 2
            
            # Save the file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Save Export File"
            )
            
            if file_path:
                wb.save(file_path)
                self.export_file_path = file_path
                self.session_exported = True
                self.session_active = False
                self.prompt_shown = False
                self.session_status_var.set("Session exported successfully")
                messagebox.showinfo("Success", "Data exported successfully")
                
                # Clear the session after successful export
                self.current_session = []
                self.session_status_var.set("No active session")
                self._toggle_input_fields(False)
                self._show_session_prompt()
                
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {str(e)}")

    def _validate_alpha_input(self, value):
        """Validate name input to accept only alphabetic characters and spaces"""
        # Allow empty strings (for field clearing)
        if value == "":
            return True
        
        # Check if the input string contains only alphabetic characters and spaces
        return all(c.isalpha() or c.isspace() for c in value)

    def _validate_numeric_input(self, value):
        """Validate oracle number input to accept only numeric characters"""
        # Allow empty strings (for field clearing)
        if value == "":
            return True
        
        # Check if the input string contains only numeric characters
        return all(c.isdigit() for c in value)
            
    def _toggle_results(self):
        """Toggle the visibility of the results section in a popup window"""
        if self.results_popup is not None and self.results_popup.winfo_exists():
            self.results_popup.destroy()
            self.results_popup = None
            self.show_results_button.config(text="Show Results")
        else:
            self._show_results_popup()
            self.show_results_button.config(text="Hide Results")

    def _show_results_popup(self):
        """Create and show the results popup window"""
        # Create a new top-level window
        self.results_popup = tk.Toplevel(self.root)
        self.results_popup.title("Salary Progression Results")
        self.results_popup.geometry("800x600")
        
        # Make the popup window modal
        self.results_popup.transient(self.root)
        self.results_popup.grab_set()
        
        # Add a close button
        close_button = ttk.Button(self.results_popup, text="Close", command=self._close_results_popup)
        close_button.pack(pady=5)
        
        # Create the Treeview in the popup window
        columns = ('date', 'event', 'grade', 'step', 'salary')
        results_tree = ttk.Treeview(self.results_popup, columns=columns, show='headings', height=25)
        results_tree.heading('date', text='Date')
        results_tree.heading('event', text='Event')
        results_tree.heading('grade', text='Grade')
        results_tree.heading('step', text='Step')
        results_tree.heading('salary', text='Salary')

        # Adjust column widths for better visibility
        results_tree.column('date', width=100)
        results_tree.column('event', width=150)
        results_tree.column('grade', width=80)
        results_tree.column('step', width=80)
        results_tree.column('salary', width=120)

        results_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Add scrollbar
        scrollbar = ttk.Scrollbar(self.results_popup, orient=tk.VERTICAL, command=results_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        results_tree.configure(yscrollcommand=scrollbar.set)

        # Display progression data if available
        if hasattr(self, 'progression_data'):
            for record in self.progression_data:
                salary_value = record['salary'] if record['salary'] is not None else 0.0
                formatted_salary = f"{salary_value:.2f}"
                results_tree.insert('', 'end', values=(
                    DateHandler.format_date(record['date']),
                    record['event'],
                    record['grade'],
                    record['step'],
                    formatted_salary
                ))

        # Store the treeview reference
        self.popup_results_tree = results_tree

        # Handle window close
        self.results_popup.protocol("WM_DELETE_WINDOW", self._close_results_popup)

    def _close_results_popup(self):
        """Close the results popup window"""
        if self.results_popup is not None:
            self.results_popup.destroy()
            self.results_popup = None
            self.show_results_button.config(text="Show Results")

    def _close_application(self):
        """Close the application with confirmation"""
        if messagebox.askyesno("Confirm Exit", "Are you sure you want to exit the application?"):
            self.root.destroy()

    def _clear_promotions(self):
        """Clear all promotions from the history"""
        if messagebox.askyesno("Confirm Clear", "Are you sure you want to clear all promotions?"):
            for item in self.promotion_tree.get_children():
                self.promotion_tree.delete(item)
            self.promotion_list = []
            self.status_var.set("All promotions cleared")

    def _save_for_export(self):
        """Save current entry for export and refresh the form"""
        if not self._validate_session():
            return
            
        try:
            # Add current calculation to session if one is active
            if not self.session_exported:
                self._add_to_session()
            
            # Store current unit for later use
            current_unit = self.unit_var.get()
            
            # Clear all fields except unit
            self._clear_all()
            
            # Restore the unit
            self.unit_var.set(current_unit)
            
            # Show success message
            self.status_var.set("Entry saved")
            
            # Reset status message after 2 seconds
            self.root.after(2000, lambda: self.status_var.set("Ready for new entry"))
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save entry: {str(e)}")

    def _toggle_input_fields(self, enabled):
        """Enable or disable input fields based on session status"""
        # Personal Information fields are always enabled
        self.name_entry.configure(state='normal')
        self.oracle_entry.configure(state='normal')
        self.sex_dropdown.configure(state='readonly')
        self.dob_entry.configure(state='normal')
        
        # Employee Information fields
        if enabled:
            self.unit_dropdown.configure(state='readonly')
            self.subtype_dropdown.configure(state='readonly')
            self.appointment_date_var.set('')
            self.initial_grade_var.set(1)
            self.initial_step_var.set(1)
        else:
            self.unit_dropdown.configure(state='disabled')
            self.subtype_dropdown.configure(state='disabled')
        self.appointment_date_var.set('')
        self.initial_grade_var.set(1)
        self.initial_step_var.set(1)
        
        # Promotion fields - only enabled when session is active
        if enabled and self.session_active:
            self.promotion_date_entry.configure(state='normal')
            self.promotion_grade_combo.configure(state='readonly')
            self.promotion_type_combo.configure(state='readonly')
            self.add_promotion_button.configure(state='normal')
        else:
            self.promotion_date_entry.configure(state='disabled')
            self.promotion_grade_combo.configure(state='disabled')
            self.promotion_type_combo.configure(state='disabled')
            self.add_promotion_button.configure(state='disabled')
        self.promotion_date_var.set('')
        self.promotion_grade_var.set('')
        self.promotion_type_var.set('Promotion')
        
        # Clear promotion list and tree
        for item in self.promotion_tree.get_children():
            self.promotion_tree.delete(item)
        self.promotion_list = []
        
        # Clear final status
        self.final_status_var.set('')
        
        # Enable/disable all input widgets except personal information
        widgets_to_toggle = [
            self.unit_var, self.subtype_var, self.appointment_date_var,
            self.initial_grade_var, self.initial_step_var,
            self.promotion_date_entry, self.promotion_grade_combo,
            self.promotion_type_combo, self.add_promotion_button
        ]
        
        for widget in widgets_to_toggle:
            if hasattr(widget, 'configure'):
                widget.configure(state='normal' if enabled else 'disabled')
            elif isinstance(widget, tk.StringVar) or isinstance(widget, tk.IntVar):
                widget.set('')

    def _show_session_prompt(self):
        """Show initial session prompt"""
        # No longer show the session prompt
        self.session_status_var.set("Ready for entries")

    def _prompt_new_session(self, event=None):
        """Prompt user to start a new session when attempting to use promotion fields"""
        if not self.session_active:
            if messagebox.askyesno("Start New Session", 
                                 "You need to start a new session before adding promotions.\n\n"
                                 "Would you like to start a new session now?"):
                self._new_session()
            return "break"  # Prevent the event from propagating
        return None

def main():
    """Main entry point"""
    try:
        root = tk.Tk()
        app = SalaryProgressionApp(root)
        root.mainloop()
    except Exception as e:
        print(f"Application error: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()